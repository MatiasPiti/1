"""Carga masiva inicial desde Excel (.xlsx) o CSV.

Columnas esperadas (encabezados case-insensitive, en español):
Código | Nombre | Precio Venta | Stock Inicial | Proveedor
"""

import csv
import os
import uuid
from datetime import datetime

from pos_core.db import transaction

_ALIASES = {
    "codigo": {"codigo", "código", "code", "sku"},
    "nombre": {"nombre", "producto", "descripcion", "descripción"},
    "precio_venta": {"precio venta", "precio_venta", "preciodeventa", "precio"},
    "stock_inicial": {"stock inicial", "stock_inicial", "stock", "cantidad"},
    "proveedor": {"proveedor", "supplier"},
}


def _mapear_columnas(encabezados: list) -> dict:
    normalizados = [h.strip().lower() for h in encabezados]
    mapeo = {}
    for campo, alias_set in _ALIASES.items():
        for idx, h in enumerate(normalizados):
            if h in alias_set:
                mapeo[campo] = idx
                break
    faltantes = {"codigo", "nombre", "precio_venta", "stock_inicial"} - mapeo.keys()
    if faltantes:
        raise ValueError(f"Faltan columnas obligatorias en el archivo: {sorted(faltantes)}")
    return mapeo


def _leer_filas(ruta: str):
    ext = os.path.splitext(ruta)[1].lower()
    if ext == ".csv":
        with open(ruta, newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            filas = list(reader)
    elif ext in (".xlsx", ".xlsm"):
        import openpyxl  # import perezoso
        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        ws = wb.active
        filas = [[("" if c is None else c) for c in fila] for fila in ws.iter_rows(values_only=True)]
    else:
        raise ValueError(f"Formato no soportado: {ext} (usar .xlsx o .csv)")
    if not filas:
        raise ValueError("El archivo está vacío")
    return filas


def cargar_masivo(ruta: str, *, usuario: str, origen: str = "MAESTRO") -> dict:
    filas = _leer_filas(ruta)
    encabezados, *datos = filas
    mapeo = _mapear_columnas([str(h) for h in encabezados])

    creados, actualizados, errores = 0, 0, []
    now = datetime.now().isoformat(timespec="milliseconds")
    sincronizado = 1 if origen == "MAESTRO" else 0

    for n, fila in enumerate(datos, start=2):
        if not fila or all(str(c).strip() == "" for c in fila):
            continue
        try:
            codigo = str(fila[mapeo["codigo"]]).strip()
            nombre = str(fila[mapeo["nombre"]]).strip()
            precio = float(fila[mapeo["precio_venta"]] or 0)
            stock = int(float(fila[mapeo["stock_inicial"]] or 0))
            proveedor = str(fila[mapeo["proveedor"]]).strip() if "proveedor" in mapeo else None
            if not codigo or not nombre:
                raise ValueError("código o nombre vacío")

            with transaction() as conn:
                existente = conn.execute(
                    "SELECT id FROM Productos WHERE codigo = ?", (codigo,)
                ).fetchone()
                if existente:
                    conn.execute(
                        """UPDATE Productos SET nombre=?, precio_venta=?, proveedor=?,
                           actualizado_en=?, version=version+1 WHERE codigo=?""",
                        (nombre, precio, proveedor, now, codigo),
                    )
                    actualizados += 1
                else:
                    conn.execute(
                        """INSERT INTO Productos
                           (uuid_unico, codigo, nombre, precio_venta, stock, proveedor, origen, sincronizado)
                           VALUES (?,?,?,?,?,?,?,?)""",
                        (str(uuid.uuid4()), codigo, nombre, precio, stock, proveedor, origen, sincronizado),
                    )
                    conn.execute(
                        """INSERT INTO Movimientos_Stock
                           (uuid_unico, producto_codigo, tipo, cantidad, stock_resultante,
                            motivo, usuario, origen, fecha_hora, sincronizado)
                           VALUES (?,?,?,?,?,?,?,?,?,?)""",
                        (str(uuid.uuid4()), codigo, "ENTRADA_EXCEL", stock, stock,
                         "Carga masiva inicial", usuario, origen, now, sincronizado),
                    )
                    creados += 1
        except Exception as e:
            errores.append({"fila": n, "error": str(e)})

    return {"creados": creados, "actualizados": actualizados, "errores": errores}


def exportar_lista_precios(ruta: str) -> int:
    """Exporta toda la tabla Productos a un .xlsx con las mismas columnas
    que espera cargar_masivo() (Código | Nombre | Precio Venta | Stock
    Inicial | Proveedor), para poder llevar ese mismo archivo a un USB de
    emergencia y volver a cargarlo desde "Carga Excel": como el código ya
    existe ahí, cargar_masivo() lo toma como actualización de precio en
    vez de crear un producto duplicado.

    Devuelve la cantidad de productos exportados.
    """
    import openpyxl
    from pos_core.db import get_connection

    conn = get_connection()
    rows = conn.execute(
        "SELECT codigo, nombre, precio_venta, stock, proveedor FROM Productos ORDER BY nombre"
    ).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"
    ws.append(["Código", "Nombre", "Precio Venta", "Stock Inicial", "Proveedor"])
    for r in rows:
        ws.append([r["codigo"], r["nombre"], r["precio_venta"], r["stock"], r["proveedor"] or ""])
    wb.save(ruta)
    return len(rows)
